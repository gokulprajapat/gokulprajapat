# updated on 07-06-2024
import time
from selenium.common.exceptions import ElementClickInterceptedException
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver import Keys
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import gspread
import pandas as pd
from google.auth import credentials
from gspread.utils import rowcol_to_a1
import re



class cis:       
    # CONSTRUCTOR FUNCTION   
    def __init__(self):  

        self.driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        # self.driver.get("http://172.18.224.253/swecourtis/index.php")
        # self.driver.maximize_window()
        self.wait= WebDriverWait(self.driver,12,3,ignored_exceptions=[ElementClickInterceptedException])

        # Authenticate and open the spreadsheet
        self.gc = gspread.service_account(filename='/opt/lampp/htdocs/php/credentials.json')
              
        
    def login(self):
        self.driver.get("http://172.18.224.253/swecourtis/index.php")
        self.driver.maximize_window()
        # login
        select = Select(self.driver.find_element(By.NAME, 'databasetype'))
        select.select_by_index(1)
        # time.sleep(1)
        user=self.driver.find_element(By.ID,"username")
        user.send_keys("supuser")
        passw=self.driver.find_element(By.ID,"pass_word")
        passw.send_keys("supuser")        
        self.wait.until(EC.element_to_be_clickable((By.ID, "login1"))).click()
        # element = self.driver.find_element(By.ID, "login1")
        # element.click()
        # time.sleep(1)
    # showing left side panel 
        # self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='show_me']"))).click()
        self.show_side_panel()
        # login close

    # for scrutiny of files
    def objection(self):
        self.driver.find_element(By.XPATH,"//span[normalize-space()='Registration Section']").click()
        self.driver.find_element(By.XPATH,"//span[normalize-space()='Case Scrutiny']").click()
        self.driver.find_element(By.XPATH,"//span[normalize-space()='Case Objection']").click()
        
        # switch to frame   
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))   

        # fileread= gc.open("CIS_UPDATED_FILE").worksheet("filling_output")
        worksheet=self.gc.open("CIS_UPDATED_FILE").worksheet("filling_output")
        # fetch all records from gsheet in form of dictionary
        dictionary=worksheet.get_all_records()

        for r in dictionary:
            file=r["FILLING_NO"]
            # print(file)
            # self.wait.until(EC.text_to_be_present_in_element((By.XPATH,"//select[@id='ffiling_no']"),file))
            time.sleep(1)
            self.wait.until(EC.presence_of_element_located((By.XPATH,"//select[@id='ffiling_no']"))).send_keys(file)
            self.wait.until(EC.element_to_be_clickable((By.ID,"button1"))).click()
            # self.driver.find_element(By.ID,"button1").click()
        
        
        self.driver.switch_to.default_content()
        
    def date_modify(self):
        # wb=openpyxl.load_workbook("date_modification.xlsx")
        worksheet = self.gc.open("CIS_UPDATED_FILE").worksheet("date_modification")
        
        # print(sheet.max_row)
        # print(sheet.max_column)
        self.driver.find_element(By.XPATH,"//span[normalize-space()='Filing Counter']").click()
        self.wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Case and Caveat Fili"))).click()
    # edit
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//img[@title='Edit']"))).click()
    # switch to frame   
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))

        for r in worksheet.get_all_records():
            case_type=r["CIVIL / CRIMINAL"]
            file_name=r["CASE_TYPE"]
            file_no=r["NUMBER"]
            file_year=r["FILLING_YEAR"]

        # selecting case type civil/criminal
            if case_type == 'Civil':
                self.driver.find_element(By.XPATH,"//input[@id='ftype_of_filing_case']").click()
            else:
                self.driver.find_element(By.XPATH,"//input[@id='ftype_of_filing_criminal']").click()
            # print(case_type,file_name,file_no)
            time.sleep(1)
        # selecting file_type   
            self.driver.find_element(By.XPATH,"//select[@id='ffiling_no_type']").send_keys(file_name)
        # filing no
            self.driver.find_element(By.XPATH,"//input[@id='ffiling_no']").clear()
            self.driver.find_element(By.XPATH,"//input[@id='ffiling_no']").send_keys(file_no)
        # filing year
            # self.driver.find_element(By.XPATH,"//input[@id='ffiling_no_year']").send_keys(file_year)
        # GO
            self.driver.find_element(By.XPATH,"//input[@id='go']").click()
            self.driver.find_element(By.XPATH,"//li[@id='tab2']").click()
        # change date
            self.driver.find_element(By.XPATH,"//input[@id='fchange_filing_date']").click()
            self.driver.find_element(By.XPATH,"//input[@id='fchange_filing_date']").click()
        # submit
            self.driver.find_element(By.XPATH,"//input[@id='submitdata']").click()
            
        # switch to parent frame
        self.driver.switch_to.default_content()

    def registration(self):        
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Registration Section']"))).click()
        self.wait.until(EC.element_to_be_clickable((By.LINK_TEXT,"Case Registration"))).click()
   

        # switch to frame   
        # self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))
        self.wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"//iframe[@id='ifr']")))
        
     #cleaning registration output file
        # wb =Workbook()
        # wb.save("registration_output.xlsx")

        # file no 
        # fileread= openpyxl.load_workbook("filling_output.xlsx")  
        # sheetread=fileread.active
        worksheet = self.gc.open("CIS_UPDATED_FILE").worksheet("filling_output")
        
        for r in worksheet.get_all_records():
            file=r["FILLING_NO"]
            first_party=r["FIRST_PARTY"]
            first_party_hin=r["FIRST_PARTY_HIN"]
            adv_bar_no=r["ADV_BAR_NO"]
            second_party=r["SECOND_PARTY"]
            second_party_hin=r["SECOND_PARTY_HIN"]
            first_add=r["ADD_FIRST"]
            first_add_hin=r["ADD_FIRST_HIN"]
            second_add=r["ADD_SECOND"]
            second_add_hin=r["ADD_SEC_HIN"]
            file_code=r["FILE_CODE"]
            next_date=r["NEXT_DATE"]
            purpose=r["PURPOSE_LISTING"]
            court=r["TRANSFER_COURT"]

            # time.sleep(2)
            # self.driver.find_element(By.XPATH,"//select[@id='ffiling_no']").send_keys(file)

            # REGISTRATION START SELECTING FILE
            select= Select(self.wait.until(EC.visibility_of_element_located((By.XPATH,"//select[@id='ffiling_no']"))))
            # print(file) 
            select.select_by_visible_text(file)
            while True:
             st=self.driver.execute_script("return document.querySelector('#fpet_name').value")
             if(len(st)>0):
              break
            
        
        # ENTERING FIRST PARTY ADDRESS
            self.driver.find_element(By.XPATH,"//textarea[@id='fpetadd']").clear()
            self.driver.find_element(By.XPATH,"//textarea[@id='fpetadd']").send_keys(first_add)

            self.driver.find_element(By.XPATH,"//textarea[@id='flpetadd']").clear()
            self.driver.find_element(By.XPATH,"//textarea[@id='flpetadd']").send_keys(first_add_hin)
            # first party name in hindi
            self.driver.find_element(By.XPATH,"//input[@id='flpet_name']").clear()
            self.driver.find_element(By.XPATH,"//input[@id='flpet_name']").send_keys(first_party_hin)

          # entering advocate barcode check for barcode reg
            if re.search("[^a-zA-Z\s]", adv_bar_no):
           #  select all             
             self.driver.execute_script("document.querySelector('#fadv_type1').options[1].selected='selected'")            
            #  press on reset button
             self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='reset_adv1']"))).click()            
             self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_name1']"))).clear()
             self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_name1']"))).send_keys(adv_bar_no)
             time.sleep(1)
            # selecting from drop down advocate
             self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_name1']"))).send_keys(Keys.ARROW_DOWN)
             self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_name1']"))).send_keys(Keys.ENTER)            
            # enter submit
            self.driver.find_element(By.XPATH,"//input[@id='pet_submit']").click()
            time.sleep(2)

            # entering advocate of 2nd party
            if second_party.upper().find("STATE OF RAJASTHAN") != -1:
                self.driver.find_element(By.ID,"fadv_name2").clear()
                self.driver.find_element(By.ID,"fadv_name2").send_keys("PP")

            # enteing 2nd party address 
            self.driver.find_element(By.XPATH,"//textarea[@id='fresadd']").clear()
            self.driver.find_element(By.XPATH,"//textarea[@id='fresadd']").send_keys(second_add)

            # second party name in hindi            
            self.driver.find_element(By.XPATH,"//input[@id='flres_name']").clear()
            self.driver.find_element(By.XPATH,"//input[@id='flres_name']").send_keys(second_party_hin)

            # second party address hindi
            self.driver.find_element(By.XPATH,"//textarea[@id='flresadd']").clear()       
            self.driver.find_element(By.XPATH,"//textarea[@id='flresadd']").send_keys(second_add_hin)
            
            self.driver.find_element(By.XPATH,"//input[@id='res_submit']").click()
            time.sleep(1)

            # jump to next date tab
            self.driver.find_element(By.XPATH,"//li[@id='reg_tab']").click()
        # select nature
            select =self.driver.find_element(By.ID,"fnaturecode_s1")
            option=Select(select)
            option.select_by_index(file_code)
            # time.sleep(1)

            # selecting next hearing date
            self.driver.find_element(By.XPATH,"//input[@id='flisting_date']").send_keys(next_date)
            # time.sleep(1)
            # for waiting purpose
            # self.wait.until(EC.presence_of_element_located((By.XPATH,"//option[@value='553']")))
            # entering purpose of listing
            select= self.wait.until(EC.visibility_of_element_located((By.XPATH,"//select[@id='fpurpose_codes']")))
            option=Select(select)        
            option.select_by_visible_text(purpose)
            # time.sleep(2)    
            # submit
            self.driver.find_element(By.XPATH,"//input[@id='register_submit']").click()
            time.sleep(1)
            # registration no extract
            reg=self.driver.find_element(By.ID,"info1").text
            x=reg.index('-')+1
            print(reg[x:len(reg)])
            print(first_party," v/s ",second_party)

            reg_no=reg[x:len(reg)]
            data=[reg_no,first_party,second_party,court]
            # insert data into google sheet
            self.gc.open("CIS_UPDATED_FILE").worksheet("registration_output").append_row(data)
            # ls.append(data)


       
        # back to main frame
        self.driver.switch_to.default_content()

    def allocation(self):
        # wb=openpyxl.load_workbook("registration_output.xlsx")
        # ws=wb.active

        worksheet = self.gc.open("CIS_UPDATED_FILE").worksheet("registration_output")

        self.driver.find_element(By.XPATH,"//span[normalize-space()='Case Allocation']").click()
        # INDIVIDUAL        
        # self.driver.find_element(By.XPATH,"//a[@id='213']").click()
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@id='213']"))).click()

        # switching to sub frame        
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))   
        # time.sleep(1)

        for r in worksheet.get_all_records():
            file=r["REGISTRATION_NO"]

            # selecting list cases
            # WAIT
            self.wait.until(EC.visibility_of_element_located((By.TAG_NAME,"body")))
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='listcase']"))).click()
            time.sleep(1)
            select=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@id='freg_no']")))
            option=Select(select)
            option.select_by_visible_text(file)
        # selecting dj court
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='1']"))).click()
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='submitdata']"))).click()
    # back to parent frame
        self.driver.switch_to.default_content()

    def show_side_panel(self):        
    # switch to main panel
     self.driver.switch_to.default_content()
    # showing left side panel 
     element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='show_me']")))
     if element.get_property('value') == ">>":
      element.click()

# establish transfer function for WAC transfer
    def establish_transfer(self,file):

        self.driver.switch_to.default_content()
        
        # click on establish transfer
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Establishment Transfer']"))).click()        




    # switching to sub frame
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))    
        # print("switching to sub frame")

        # selecting dj
        self.wait.until(EC.presence_of_element_located((By.TAG_NAME,"body")))#waiting some time
        select=self.wait.until(EC.presence_of_element_located((By.XPATH,"//select[@id='fsource_est_name']")))
        option=Select(select)
        option.select_by_visible_text("DJ/ADJ/Civil/Criminal Cases-1")

        # selecting file type
        select=self.wait.until(EC.presence_of_element_located((By.XPATH,"//select[@id='fsourcecase_type']")))
        # typing case type
        select.send_keys(file[0])
        # type registration no         
        self.wait.until(EC.presence_of_element_located((By.XPATH,"//input[@id='fsource_case_no']"))).send_keys(file[1])
        # type year 
        self.wait.until(EC.presence_of_element_located((By.XPATH,"//input[@id='fsourceyear']"))).send_keys(file[2])
        # press go
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='go']"))).click()
        # waiting some time
        self.wait.until(EC.visibility_of_element_located((By.XPATH,"//select[@id='ftarget_est_name']")))

        # select nature of disposal
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@id='flegacy_disptype']"))).send_keys("Transferred-563")
        # select uncontested
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='uncontested']"))).click()

        # selecting file type
        select=self.wait.until(EC.presence_of_element_located((By.XPATH,"//select[@id='ftarget_est_name']")))
        # typing establishment name
        select.send_keys("Wac Cases-7")
        
        # selecting case type
        select=self.wait.until(EC.presence_of_element_located((By.XPATH,"//select[@id='ftargetcase_type']")))

        # typing file type
        select.send_keys(file[0])
        # selecting register        waiting some time
        self.wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='showregister']")))
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='showregister']"))).click()

    # wac registration no   waiting some time
        self.wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='ftarget_reg_no']")))
        reg_no=self.wait.until(EC.presence_of_element_located((By.XPATH,"//input[@id='ftarget_reg_no']"))).get_property('value')

    # submit     waiting some time
        self.wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='insert']")))
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='insert']"))).click()
        
        # wb=openpyxl.load_workbook('wac.xlsx')
        # ws=wb.active
        wac_reg=str(file[0])+"/"+str(reg_no)+"/"+str(file[2])
        # update wac google spread sheet
        self.gc.open("CIS_UPDATED_FILE").worksheet("wac").append_row([wac_reg,"/".join(file)])
        print("WAC Reg. No. ",wac_reg," of ","/".join(file))
        
     

    # BACK TO PARENT FRAME
        self.driver.switch_to.default_content()
        
    # click on individual allocation
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Individual Case']"))).click()
    # switching to sub frame
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))





    def file_transfer(self):        
        # click on admin menu
        self.driver.find_element(By.XPATH,"//span[normalize-space()='Admin Menu']").click()
        # transfer of case
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Transfer of Case']"))).click()
        # click on individual allocation
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Individual Case']"))).click()

    # switching to sub frame
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))

     
        ws = self.gc.open("CIS_UPDATED_FILE").worksheet("registration_output")

        for r in ws.get_all_records():
            # extracting file type
            file=r["REGISTRATION_NO"].split('/')

            court=r["TRANSFER_COURT"]

            if court == "NOT":
                continue
            elif court == "WAC":
                self.establish_transfer(file)
                continue
         

        # hard explicit wait added here
            # time.sleep(1)
            self.wait.until(EC.visibility_of_element_located((By.TAG_NAME,'body')))
            self.wait.until(EC.visibility_of_element_located((By.XPATH,"//td[@id='formtable']//table")))
            self.wait.until(EC.visibility_of_element_located((By.XPATH,"//option[@value='527']")))

        #   selecting case type
            select=self.wait.until(EC.visibility_of_element_located((By.XPATH,"//select[@id='fmm_case_type']")))
    # selecting file
            for item in select.find_elements(By.TAG_NAME,"option"):
             if item.text.find(file[0]) != -1:
              option=Select(select)
              while True:
               option.select_by_visible_text(item.text)
               if option.first_selected_option.text == "Select":
                continue
               else:
                break   #exit from inner loop
              break   #exit from outer loop
             
            #  case no
            self.wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='fmm_case_no']"))).send_keys(file[1])
            # case year
            # print(file[2])
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fmm_case_year']"))).send_keys(file[2])
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='mm_search']"))).click()

            select=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@id='fcourt_no_s']")))
            
    # selecting court
            for item in select.find_elements(By.TAG_NAME,'option'):
                if item.text.find(str(court)) != -1:
                    option=Select(select)
                    option.select_by_visible_text(court)
                    break
           
            current_time = datetime.now()
            year=str(current_time.year)

            if len(str(current_time.month))==1:
                month="0"+str(current_time.month)
            else:
                month=str(current_time.month)

            if len(str(current_time.day))==1:
                day="0"+str(current_time.day)
            else:
                day=str(current_time.day)

            dt=day+"-"+month+"-"+year
            # entering transfer date or registration date
            self.driver.find_element(By.XPATH,"//input[@id='ftransfer_date']").send_keys(dt)
            # enter submit
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='submitdata']"))).click()

        
        # BACK TO PARENT FRAME
        self.driver.switch_to.default_content()



    def filling(self):
        

        s=self.gc.open("CIS_UPDATED_FILE").worksheet("complete_registration")
        

        # clicking filling counter
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Filing Counter']"))).click()
        # clicking case and caveat 
        self.wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT,"Case and Caveat Fili"))).click()
        # switch to frame   
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))             
        # time.sleep(1)

    

        for row in s.get_all_records(numericise_ignore=["all"]):
            case_type=row["CASE_TYPE"]
            file_type=row["CIVIL / CRIMINAL"]
            first_party=row["FIRST_PARTY"]
            first_party_hindi=row["FIRST_PARTY_HINDI"]            
            first_party_age=row["AGE"]
            first_party_gender=row["FIRST_GENDER"]
            adv_bar_no=row["ADV_BAR_NO"]
            advocate=row["ADVOCATE"]
            advocate_hindi=row["ADVOCATE_HINDI"]
            second_party=row["SECOND_PARTY"]
            second_party_hindi=row["SECOND_PARTY_HINDI"]
            second_party_gender=row["SECOND_GENDER"]
            police_station=row["POLICE_STATION"]
            fir_number=row["FIR_NUMBER"]
            fir_year=row["YEAR_OF_FIR"]
            act=row["ACT"]
            section=row["SECTION"]
            first_party_add=row["ADDRESS 1"]
            first_party_add_hindi=row["ADDRESS 1 HINDI"]
            second_party_add=row["ADDRESS 2"]
            second_party_add_hindi=row["ADDRESS 2 HINDI"]
            organisation_first=row["ORGANISATION FIRST"]
            organisation_second=row["ORGANISATION SECOND"]
            file_code=row["FILE CODE"]
            next_date=row["NEXT DATE"]
            purpose=row["PURPOSE"]
            court=row["TRANSFER_COURT"]

            
        # according to civil or criminal file select radio button
            if case_type == 'Civil':
                time.sleep(1)
                self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ftype_of_filing_case']"))).click()
            else:              
                time.sleep(1)
                self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='ftype_of_filing_criminal']"))).click()

            time.sleep(1)
            select=Select(self.wait.until(EC.element_to_be_clickable((By.NAME, "ffiling_no_type"))))
            
            # selecting file type
            print(file_type)
            select.select_by_visible_text(file_type)

            # file type
            self.wait.until(EC.element_to_be_clickable((By.ID,"tab1"))).click()

            # uncheck checkbox
            if organisation_first == 'ORGANISATION' and case_type == 'Civil':        
                self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fpetorg_check']"))).click()                

            if organisation_first != 'ORGANISATION' and case_type == 'Criminal':        
                self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fpetorg_check']"))).click()                

        # first party name
            if organisation_first != 'ORGANISATION':        
                self.wait.until(EC.presence_of_element_located((By.XPATH,"//input[@id='fpet_name']"))).send_keys(first_party)
            else:
                self.wait.until(EC.presence_of_element_located((By.XPATH,"//input[@id='fpetorg_type']"))).send_keys(first_party)

        # first party age            
            if organisation_first != 'ORGANISATION':        
                self.wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@id='fpet_age']"))).send_keys(first_party_age)

        # first party gender
            if organisation_first != 'ORGANISATION' and first_party_gender.upper() == "FEMALE":        
                self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fpet_sex_female']"))).click()

        # advocate name
            self.driver.find_element(By.XPATH,"//input[@id='fadv_name1']").send_keys(advocate)
       
        # organisation or not second party second party detail
            if organisation_second == 'ORGANISATION':       
                self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fresorg_check']"))).click()
                time.sleep(1) 
                # second party name
                self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fresorg_type']"))).send_keys(second_party)
            else:
                if second_party_gender.upper() == 'FEMALE':
                    self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fres_sex_female']"))).click()
                self.driver.find_element(By.XPATH,"//input[@id='fres_name']").send_keys(second_party)


        # first party hindi name  
            self.driver.find_element(By.XPATH,"//input[@id='flpet_name']").send_keys(first_party_hindi)
        # second party hindi name
            self.driver.find_element(By.XPATH,"//input[@id='flres_name']").send_keys(second_party_hindi)
        # advocate hindi name
            self.driver.find_element(By.ID,"fladv_name1").send_keys(advocate_hindi)
            # next page
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='next1']"))).click()

        # fir and police station 
            if case_type == 'Criminal' and police_station != "NOT":
                select=Select(self.driver.find_element(By.XPATH,"//select[@id='fpolice_st_code']"))
                select.select_by_visible_text(police_station)
                self.driver.find_element(By.XPATH,"//input[@id='ffir_no']").send_keys(fir_number)
                self.driver.find_element(By.XPATH,"//input[@id='ffir_year']").send_keys(fir_year)

            
        # act name
            
            acts=str(act).split("\\")
            
            sections=str(section).split("\\")
            counter=True
            # for more act and section increment
            t=1
            for (a,s) in zip(acts,sections):
                # domestic violence hide party
                if a == "DOMESTIC VIOLENCE ACT 2005-1099" and counter == True:
                    # hide party name
                    self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fhide_partyname']"))).click() 
                    counter=False
       
                time.sleep(1)
 
                self.driver.find_element(By.XPATH,f"//input[@id='fdispactcodes{t}']").send_keys(a[0:len(a)-5])
                      
                time.sleep(1)
                self.driver.find_element(By.ID,"fdispactcodes"+str(t)).send_keys(Keys.ARROW_DOWN)

                self.driver.find_element(By.ID,"fdispactcodes"+str(t)).send_keys(Keys.ENTER)

                
            # section number
                self.driver.find_element(By.ID,"factsection_code"+str(t)).send_keys(s)
            # click more act section            
  
                try:
                    self.driver.find_element(By.XPATH,f"//input[@id='moreacts{t}']").click()
                except:
                    self.driver.find_element(By.XPATH,f"//td[@align='right']//input[@id='moreacts{t}']").click()
    
                t=t+1
                
            
                
        # submit
            self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='submitdata']"))).click()
            
            #filling no exract
            element=self.driver.find_element(By.XPATH,"//span[@id='show_filingno']")
            

            # until not visible filling no
            while len(element.text) == 0:                
                element=self.driver.find_element(By.XPATH,"//span[@id='show_filingno']")

            
            # trim=file_type[0:len(file_type)-4]

            print(first_party,"v/s",second_party,"\n",element.text)
            # index=element.text.index(trim)
            # finding 2024 index
            # last=element.text.index("/2024")+5

            # filling for extracting 
            x=element.text.split("\n")#this line newly added
            filling_no=x[1][13:]#this line newly added
            cnr_no=x[2][6:]#this line newly added
            # filling_no=element.text[index:last]
            # index=element.text.index("RJJ")

            # cnr_no=element.text[index:index+16]

            data=[filling_no,cnr_no,first_party,first_party_hindi,adv_bar_no    ,second_party,second_party_hindi,first_party_add,first_party_add_hindi,second_party_add,second_party_add_hindi,file_code,next_date,purpose,court]            
            self.gc.open("CIS_UPDATED_FILE").worksheet("filling_output").append_row(data)

        # excel for writing
     
        
        

        # print("write successfully...")

    # back to default frame
        self.driver.switch_to.default_content()

    def date_change(self,dt1):
        self.driver.switch_to.default_content()
        # clicking on case proceding
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id='cssmenu']/ul/li[13]/a/span"))).click()
    # clicking on select date
        self.wait.until(EC.element_to_be_clickable((By.ID,"183"))).click()

    # switching to sub frame
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))

    #select date
        element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fdate']")))
        element.clear()
        element.send_keys(dt1)

    def holiday_check(self, dt):
    #  print("Holiday check")

    #  self.date_change(dt1)    
       
        # selecting establishment
     element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"(//select[@id='dbestid'])[1]")))
     options=Select(element)
     establishments=len(options.options)
        
        # CONNECTING GOOGLE SHEET CONNECTION
        # gc = gspread.service_account(filename='credentials.json')
        # worksheet = self.gc.open("holiday_date_modify").worksheet("main_file")

     workbook=Workbook()
     worksheet=workbook.active

     for r in range(1,establishments):
          # switching to parent frame
      self.driver.switch_to.default_content()
            # selecting establishment one by one
      element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"(//select[@id='dbestid'])[1]")))
      options=Select(element)
      options.select_by_index(r)

            # for entering in google sheet
            # est=options.first_selected_option.text
      est=self.driver.execute_script('''let e=document.querySelector('#dbestid');
            return e.options[e.selectedIndex].text;''')

      self.show_side_panel()

            # click on query builder
      self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Query Builder']"))).click()

            # click on query builder sub menu
            # self.wait.until(EC.element_to_be_clickable((By.XPATH,"//li[@class='has-sub active']//li[1]//a[1]"))).click()
            # self.driver.execute_script('document.querySelector("\\34 35").click()')
      self.driver.execute_script('document.querySelector("body > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > ul:nth-child(1) > li:nth-child(21) > ul:nth-child(2) > li:nth-child(1) > a:nth-child(1)").click()')

            #switch to sub frame
      self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))

            # count courts
      element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@id='fcourt_no_s']")))
      options=Select(element) 
      courts=len(options.options)
        

            # trace all courts move one by one
      for i in range(1,courts):
              #click on both
       self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='both']"))).click()

                # selecting court name one by one using i
       element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@id='fcourt_no_s']")))
       options=Select(element)        
                # options.select_by_index(i)  

       script='document.querySelector("#fcourt_no_s").getElementsByTagName("option")[{}].selected="selected";'.format(i)
       self.driver.execute_script(script)

       script='return document.querySelector("#fcourt_no_s").getElementsByTagName("option")[{}].textContent;'.format(i)
       cort=self.driver.execute_script(script)

                # while(True):
                #     if  options.first_selected_option.get_attribute("textContent") == "Select":
                #         options.select_by_index(i)
                #     else:
                #         break
                        
                # self.driver.quit()
                # for entering in google sheet
                # cort=options.first_selected_option.text
                
                

       print("****************** ",options.first_selected_option.text," *****************")      
                #Showing 0 to 0 of 0 entries 

                #click on report title and write "d" in text box
                # self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='report_title']"))).send_keys("d")
       self.driver.execute_script('document.querySelector("#report_title").value="d";')

                #click on submit button
       self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='report_button']"))).click()

                #wait for processing spinner disapper
       self.wait.until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='faded']")))

                # wait for page load completely
                # self.wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                # TimeoutException occured

                #select all
       element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@name='example_length']")))
       option=Select(element)
       option.select_by_index(5)
       #wait for processing spinner disapper
       self.wait.until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='faded']")))
    #check all holiday dates one by one in dictionary
       for dt1,dt2 in dt.items():
        #clear next date previously entered
        self.wait.until(EC.element_to_be_clickable((By.ID,"search_5"))).clear()
        #enter next date
        self.wait.until(EC.element_to_be_clickable((By.ID,"search_5"))).send_keys(dt1)
        #click on search button
        self.wait.until(EC.element_to_be_clickable((By.ID,"search_0"))).click()
        #wait for processing spinner disapper
        self.wait.until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='faded']")))
        #getting all case 
        cases=self.driver.execute_script(' return document.querySelectorAll("#example > tbody > tr")')
        flag=False
        record=len(cases)
        
        for case in cases:
          # checking of no result found on next date , if no record then break loop
         if record == 1:
          resultset=case.find_element(By.TAG_NAME,'td').get_attribute("textContent")
          if resultset.strip() == "No data available in table":
           break
         c=case.find_element(By.TAG_NAME,'a')
                        # c.click()                
         print(c.get_attribute('textContent'))
                        # entering reg_no in google sheet 
         reg=c.get_attribute('textContent')
                        # check reg no is correct if not then correct it
         reg=reg.replace("Cr ","Cr. ")
         reg=reg.replace("Misc ","Misc. ")
         reg=reg.replace("Reg ","Reg. ")
         
         worksheet.append([reg,dt2,cort,est])
         workbook.save('dem.xlsx')                       
       #click on back button
       self.wait.until(EC.element_to_be_clickable((By.XPATH,"//a[@href='query_generate.php']"))).click()
         #wait for back to page
         # self.wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,"#search_0")))
     workbook.close()
        # switching to parent frame
     self.driver.switch_to.default_content()
    
    def court_change(self,crt_name):
        self.driver.switch_to.default_content()
        # clicking on case proceding
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//*[@id='cssmenu']/ul/li[13]/a/span"))).click()
        
    # clicking on select court
        # self.wait.until(EC.element_to_be_clickable((By.ID,"182"))).click()
        self.driver.execute_script('document.querySelector("body > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > ul:nth-child(1) > li:nth-child(13) > ul:nth-child(2) > li:nth-child(1) > a:nth-child(1)").click()')

    # switching to sub frame
        self.driver.switch_to.frame(self.driver.find_element(By.XPATH,"//iframe[@id='ifr']"))

    #click for selecting court
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Select']"))).click()
        time.sleep(1)
        all_court=self.driver.execute_script('return document.querySelector("#fcourt_no_s_chosen > div > ul").getElementsByTagName("li")')

        # print("search court :",crt_name[3:])

        for i in all_court:
            # court name
            c_name=i.get_attribute('textContent')
            # print(c_name)            
            if crt_name[3:] in c_name :
                # i.click()
                box=self.driver.execute_script('''return document.querySelector("input[type='text']");''')
                box.send_keys(c_name)
                box.send_keys(Keys.ARROW_DOWN)
                box.send_keys(Keys.ENTER)
                break
        self.driver.execute_script('document.querySelector("#submitdata").click()')

        # .getElementsByTagName("li")[3].textContent

        # self.wait.until(EC.visibility_of_element_located((By.XPATH,"//input[@type='text']"))).send_keys(crt_name)
        # document.querySelector("input[type='text']").value='ACJ-Cum-MM No.2'

    # enter court name for searching
        # self.wait.until(EC.element_to_be_clickable((By.XPATH,"//div[@class='chosen-search']"))).send_keys(crt_name)
        # time.sleep(10)

        # switching to parent frame
        self.driver.switch_to.default_content()

    def est_change(self,es):
        self.driver.switch_to.default_content()
        # est_names=self.driver.execute_script('return document.querySelector("#dbestid").getElementsByTagName("option")')
        element=self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@id='dbestid']")))
        opt=Select(element)
        opt.select_by_visible_text(es)
        self.show_side_panel()
        # self.driver.switch_to.default_content()

    def get_establishment(self):
        self.driver.switch_to.default_content()
        return self.driver.execute_script(' return document.querySelector("#session_court_name").textContent')
    
    def open_modify_next_date_form(self):
     self.driver.switch_to.default_content()
        # click on admin panel
     self.driver.execute_script('document.querySelector("body > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > ul:nth-child(1) > li:nth-child(5) > a:nth-child(1) > span:nth-child(1)").click()')    
        # click on proceeding modification
     self.driver.execute_script('''document.querySelector("a[id='110'] span[class='secondul']").click()''')
        # click on modify next date and purpose
        # self.driver.execute_script('document.querySelector("").click()')
     self.wait.until(EC.element_to_be_clickable((By.XPATH,'//a[@id="314"]'))).click()

        # switch to sub frame
     self.wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"//iframe[@id='ifr']")))

    def modify_next_date(self,reg,dt):#change next date of holiday cases, if form is already open
        # registration number , case , year split 
        li=reg.split('/')
        case_type=li[0]
        # print(case_type)
        r_no=li[1]
        year=li[2]
        # extract all case type options from dropdown
        # opt=self.driver.execute_script('''document.querySelector("#fmm_case_type").getElementsByTagName('option')''')
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//select[@id='fmm_case_type']"))).send_keys(case_type)
        # opt=Select(element).options
        
        

        # self.wait.until(EC.presence_of_element_located((By.TAG_NAME,'body')))
        # self.wait.until(EC.visibility_of_all_elements_located((By.XPATH,"//option[@value='527']")))
        # self.wait.until(EC.presence_of_element_located((By.XPATH,"//option[@value='527']")))

        # for i in opt:
        #     if case_type in i.get_attribute('textContent'):                
        #         print(case_type,",",i.get_attribute('textContent'))
        #         # case type select
        #         self.driver.execute_script("arguments[0].selected='selected';",i)
        #         break

        # reg no enter
        self.driver.execute_script('document.querySelector("#fmm_case_no").value={};'.format(r_no))
        # enter year
        self.driver.execute_script('document.querySelector("#fmm_case_year").value={};'.format(year))
        # click on go
        # time.sleep(3)
        self.driver.execute_script('document.querySelector("#mm_search").click()')
        # enter next hearing date
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fnext_date']"))).send_keys(dt)
        

        # self.driver.execute_script('document.querySelector("#fnext_date").value=arguments[0];',dt)
        # print(dt)
        # enter remark
        self.driver.execute_script('document.querySelector("#fcase_remark").value="holiday";')
        # click on submit button
        self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='submitdata']"))).click()
        

    def holiday_date_modify(self):
    # Authenticate and open the spreadsheet
    #  gc = gspread.service_account(filename='credentials.json')        
     workbook=openpyxl.load_workbook("dem.xlsx")
     worksheet=workbook.active
     
    #  list_of_dicts=worksheet.get_all_records(numericise_ignore=["all"])
     pre_est=None
     pre_cort=None
     first_time=True
     for r in worksheet.values:            
      # ([r['reg_no'],r['next_date'],r['court_name'],r['est_name']])
    #   r[0]=reg no ,r[1]= next date,r[2]= courtname, r[3]= est name
      reg_no=r[0]
      next_date=r[1]
      court_name=r[2]
      est_name=r[3]
      # for first time
      if first_time :
       self.est_change(est_name)
       self.court_change(court_name)
       self.open_modify_next_date_form()
       self.modify_next_date(reg_no,next_date)          
      else:#for second time
       if pre_est == est_name:
        if pre_cort == court_name:
         self.modify_next_date(reg_no,next_date)
        else:    
         self.court_change(court_name)
         self.open_modify_next_date_form()
         self.modify_next_date(reg_no,next_date)
       else:
        self.est_change(est_name)
        self.court_change(court_name)
        self.open_modify_next_date_form()
        self.modify_next_date(reg_no,next_date)
    
      pre_est=est_name        
      pre_cort=court_name
      first_time=False
    
     workbook.close()
        
    


    def add_advocate(self,advocate,advocate_hindi,bar_reg,gender,mobile):
     self.driver.switch_to.default_content()
     self.wait.until(EC.presence_of_element_located((By.TAG_NAME,"body"))).click()
    #  click on master
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Master']"))).click()
    #  click on local master
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Local Masters']"))).click()
    #  click on advocate
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//span[normalize-space()='Advocate']"))).click()    
    #  switch to sub frame
     self.wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"//iframe[@id='ifr']")))
    #  enter advocate name
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_name']"))).send_keys(advocate)
    #  enter bar reg no. of advocate
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_reg']"))).send_keys(bar_reg)
    #  enter mobile no.
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_mobile']"))).send_keys(mobile)
    #  check gender
     if "female" == gender.lower():
      self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fadv_sex_female']"))).click()

    #  enter advocate name in hindi
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='fladv_name']"))).send_keys(advocate_hindi)

    #  click on next tab 
     self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='next']"))).click()
     time.sleep(5)
    #  click on submit
    #  self.wait.until(EC.element_to_be_clickable((By.XPATH,"//input[@id='submitdata']"))).click()

    
obj= cis()
obj.login()
obj.filling()
obj.objection()
obj.registration()
obj.allocation()
# obj.file_transfer()

# obj.date_modify()

# obj.add_advocate("d","धिवक्ता का","female","R/252/2001","9828412505")

# obj.holiday_check({"30-06-2024":"01-07-2024"})
# obj.holiday_check("22-06-2024","24-06-2024")
# obj.holiday_date_modify()

# file1=["Cr. Appeal",142,2024]

# obj.establish_transfer(file1)




